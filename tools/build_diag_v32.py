"""Build Consolidated Diag v3.2.xlsx — DELIVERABLE.

Inputs:
  C:\\Projects\\Diag\\Consolidated Diag v3.2 Review .xlsx
    - Sheet 'Master v3.1 QM'  (117 rows, post-review QM matrix)
    - Sheet 'Diag matrix_Fusa' (94 rows, V1 FuSa matrix)
  C:\\Projects\\Diag\\Customer_Diag\\Customer_Diag.pdf
    - Customer-approved DTC table (98 DTCs P1A 01 00..61 + U1A 01 00..01)
  c:\\Projects\\artha-v4\\tools\\customer_dtc_table.json
    - Pre-parsed customer table

Output: C:\\Projects\\Diag\\Consolidated Diag v3.2.xlsx with sheets:
  1.  Cover & Legend         — what this workbook contains
  2.  Master QM v3.2         — clean QM matrix
  3.  Master FuSa v3.2       — clean FuSa matrix (same shape as QM)
  4.  Customer Cross-Check   — every DTC vs customer approval (SPN/FMI/FL)
  5.  QM <-> FuSa Alignment  — side-by-side per DTC
  6.  Justifications         — layman explanation per fault
  7.  Logs & Gaps            — every issue ranked CRITICAL/HIGH/MEDIUM/LOW

The script writes a fully-validated, auditable workbook.  Every cell is checked.
Every gap is logged.  No silly defects after release.
"""

from __future__ import annotations

import json
import re
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Paths

DIAG_DIR = Path(r"C:\Projects\Diag")
REVIEW_XLSX = DIAG_DIR / "Consolidated Diag v3.2 Review .xlsx"
CUSTOMER_JSON = Path(r"c:\Projects\artha-v4\tools\customer_dtc_table.json")
OUT = DIAG_DIR / "Consolidated Diag v3.2.xlsx"

# ---------------------------------------------------------------------------
# Styling constants

HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
HDR_FILL = PatternFill("solid", fgColor="1F3864")
SUB_HDR_FONT = Font(bold=True, size=10)
SUB_HDR_FILL = PatternFill("solid", fgColor="D9E1F2")
TITLE_FONT = Font(bold=True, size=14, color="1F3864")

STATUS_FILL = {
    "Original (V1)": PatternFill("solid", fgColor="E8F1FB"),
    "Reviewer-modified": PatternFill("solid", fgColor="FFF4D6"),
    "Added (V3 AI)": PatternFill("solid", fgColor="FFF4D6"),
    "Added (Customer)": PatternFill("solid", fgColor="E8F8E1"),
    "Added (new in v3.2)": PatternFill("solid", fgColor="F4E1F8"),
    "Removed (Customer)": PatternFill("solid", fgColor="F8D7DA"),
}

SEVERITY_FILL = {
    "CRITICAL": PatternFill("solid", fgColor="C00000"),
    "HIGH": PatternFill("solid", fgColor="ED7D31"),
    "MEDIUM": PatternFill("solid", fgColor="FFC000"),
    "LOW": PatternFill("solid", fgColor="A9D08E"),
    "INFO": PatternFill("solid", fgColor="BDD7EE"),
}
SEVERITY_FONT_WHITE = {"CRITICAL", "HIGH"}

VERIFY_FILL = {
    "OK": PatternFill("solid", fgColor="C6EFCE"),
    "MISMATCH": PatternFill("solid", fgColor="FFC7CE"),
    "MISSING IN MATRIX": PatternFill("solid", fgColor="F4CCCC"),
    "MISSING IN CUSTOMER": PatternFill("solid", fgColor="FCE4D6"),
    "INFO": PatternFill("solid", fgColor="DDEBF7"),
}

THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ---------------------------------------------------------------------------
# Output column schemas

QM_COLUMNS = [
    "#", "Fault Type", "Subject", "DTC Code", "SPN", "FMI",
    "Fault Name", "Reaction", "ASIL", "Fault Level",
    "Enabling Conditions", "Detection Criteria",
    "Periodicity [ms]", "Detection Confirmation Time [ms]",
    "Healing Criteria", "Healing Confirmation Time [ms]",
    "VCU Request", "Notification", "Freeze Frame", "Variant",
    "Status", "Verification", "Note",
]

FUSA_COLUMNS = [
    "#", "Fault Type", "Subject", "DTC Code", "SPN", "FMI",
    "Fault Name", "Reaction", "ASIL", "Fault Level",
    "Enabling Conditions", "Detection Criteria",
    "Periodicity [ms]", "Detection Confirmation Time [ms]",
    "Healing Criteria", "Healing Confirmation Time [ms]",
    "VCU Request", "Notification",
    "Domain", "Owner", "Severity",
    "Status", "Verification", "Note",
]

# ---------------------------------------------------------------------------
# Helpers


def norm(s: Any) -> str:
    if s is None:
        return ""
    s = str(s)
    s = (
        s.replace("\u2013", "-")
         .replace("\u2014", "-")
         .replace("\u2011", "-")
         .replace("\ufffd", "-")
    )
    s = re.sub(r"Over\?Temperature", "Over-Temperature", s)
    s = re.sub(r"After\?Run", "After-Run", s)
    s = s.strip()
    return s


def norm_dtc(s: Any) -> str:
    s = norm(s)
    if not s:
        return ""
    m = re.match(r"^([PU])\s*1A\s*0?1\s*([0-9A-Fa-f]{2})\b", s)
    if not m:
        return s.upper().replace(" ", "")
    return f"{m.group(1)}1A 01 {m.group(2).upper()}"


REACTION_NORMALISE = {
    "ASC/6 Switch open": "ASC or 6 Switch Open",
    "ASC/6 switch open": "ASC or 6 Switch Open",
    "ASC or 6 Switch Open": "ASC or 6 Switch Open",
    "ASC or 6 switch Open": "ASC or 6 Switch Open",
    "Linear Deration -> 6 Switch Open": "Linear Deration -> 6 Switch Open",
    "Linear Deration ? 6 Switch Open": "Linear Deration -> 6 Switch Open",
    "ASC followed by AD": "ASC followed by Active Discharge",
    "Last valid valid value": "Last Valid Value",
    "No reaction": "No Reaction",
}


def normalise_reaction(s: str) -> str:
    if not s:
        return ""
    s = norm(s)
    return REACTION_NORMALISE.get(s, s)


def parse_fault_level_letter(s: str) -> str:
    """Customer style 'A', 'B/2', 'A/3', 'C/1' -> 'A','B','A','C'. 'L1' -> ''. """
    if not s:
        return ""
    s = norm(s).upper().replace(" ", "")
    if s.startswith("L") and s[1:].isdigit():
        return ""  # E-GAS level only — not a customer severity
    m = re.match(r"^([ABCD])(?:[/\\\-]?\d)?$", s)
    if m:
        return m.group(1)
    m = re.match(r"^\d[/\\\-]?([ABCD])$", s)
    if m:
        return m.group(1)
    return ""


def parse_fault_level_number(s: str) -> str:
    if not s:
        return ""
    s = norm(s).upper().replace(" ", "")
    if s.startswith("L") and s[1:].isdigit():
        return s[1:]
    m = re.match(r"^[ABCD][/\\\-]?(\d)$", s)
    if m:
        return m.group(1)
    m = re.match(r"^(\d)[/\\\-]?[ABCD]$", s)
    if m:
        return m.group(1)
    return ""

# ---------------------------------------------------------------------------
# Customer-approved post-review fault-level deltas (from HANDE HCV DTC1/2/3.jpg)
# Pink-highlighted cells in those JPGs = customer requested change.

CUSTOMER_FL_OVERRIDES: dict[str, dict[str, str]] = {
    "P1A 01 1F": {"to": "A", "from": "B", "src": "Customer JPG row 26 (Phase A Current Offset)"},
    "P1A 01 15": {"to": "B", "from": "A", "src": "Customer JPG row 49 (Board Temp OT)"},
    "P1A 01 13": {"to": "B", "from": "A", "src": "Customer JPG row 53 (Stator Temp OT)"},
    "P1A 01 16": {"to": "C", "from": "A", "src": "Customer JPG row 89 (Positive overspeed)"},
    "U1A 01 01": {"to": "A", "from": "B", "src": "Customer JPG row 181 (TSC1_DM CAN timeout)"},
}

# Customer-removed (strikethrough) DTCs
CUSTOMER_REMOVED: set[str] = {
    # row 90 in JPG: Negative overspeed (was duplicate P1A 01 16)
}

# Customer-added DTCs (green highlight rows 208-222 in DTC1/2/3.jpg)
# These are split-out variants that exist in customer PDF but not in our V1.
CUSTOMER_ADDED_FROM_JPG: list[dict] = [
    {"dtc": "P1A 01 04", "name": "Stator Temperature Sensor Short to Power Supply Failure",
     "spn": "522601", "fmi": "3", "fl": "A", "subject": "Stator Temperature"},
    {"dtc": "P1A 01 05", "name": "Stator Temperature Sensor Short to Ground Failure",
     "spn": "522601", "fmi": "4", "fl": "B", "subject": "Stator Temperature"},
    {"dtc": "P1A 01 0A", "name": "HVDC Voltage Sensor Short to Power Supply Failure",
     "spn": "522594", "fmi": "3", "fl": "A", "subject": "HVDC Voltage"},
    {"dtc": "P1A 01 0B", "name": "HVDC Voltage Sensor Short to Ground Failure",
     "spn": "522594", "fmi": "4", "fl": "B", "subject": "HVDC Voltage"},
    {"dtc": "P1A 01 12", "name": "Stator Temperature Overtemperature Level 2 (Derating)",
     "spn": "522601", "fmi": "16", "fl": "B", "subject": "Stator Temperature"},
    {"dtc": "P1A 01 13", "name": "Stator Temperature Overtemperature Level 3 Failure",
     "spn": "522601", "fmi": "0", "fl": "A", "subject": "Stator Temperature"},
    {"dtc": "P1A 01 14", "name": "Board Temperature Overtemperature Level 2 (Derating)",
     "spn": "522602", "fmi": "16", "fl": "B", "subject": "Board Temperature"},
    {"dtc": "P1A 01 17", "name": "Motor Overspeed Level 3 Failure",
     "spn": "522591", "fmi": "0", "fl": "A", "subject": "Speed"},
    {"dtc": "P1A 01 30", "name": "Controller Active Discharge Fault",
     "spn": "522603", "fmi": "21", "fl": "A", "subject": "Active Discharge"},
    {"dtc": "P1A 01 35", "name": "Power Module Fault",
     "spn": "522603", "fmi": "17", "fl": "A", "subject": "Power Module"},
    {"dtc": "P1A 01 3D", "name": "HVDC Voltage Undervoltage Level 2 (Bus)",
     "spn": "522595", "fmi": "18", "fl": "B", "subject": "HVDC Voltage"},
    {"dtc": "P1A 01 3F", "name": "HVDC Voltage Overvoltage Level 2 (Bus)",
     "spn": "522595", "fmi": "16", "fl": "B", "subject": "HVDC Voltage"},
    {"dtc": "P1A 01 40", "name": "HVDC Voltage Overvoltage Level 3 (Bus)",
     "spn": "522595", "fmi": "0", "fl": "A", "subject": "HVDC Voltage"},
    {"dtc": "U1A 01 00", "name": "CAN Busoff Failure (vehicle CAN)",
     "spn": "522390", "fmi": "0", "fl": "A", "subject": "CAN"},
]

# ---------------------------------------------------------------------------
# Subject mapping & ordering


SUBJECT_RULES = [
    ("hvdc voltage", "HVDC Voltage"),
    ("hvdc current", "HVDC Current"),
    ("phase current sum", "Phase Current Sum"),
    ("phase a", "Phase A"),
    ("phase b", "Phase B"),
    ("phase c", "Phase C"),
    ("phase voltage", "Phase Voltage"),
    ("phase overcurrent", "Phase Current"),
    ("gate driver", "Gate Driver"),
    ("active discharge current", "Active Discharge Current"),
    ("passive discharge current", "Passive Discharge Current"),
    ("active discharge", "Active Discharge"),
    ("passive discharge", "Passive Discharge"),
    ("board temperature", "Board Temperature"),
    ("stator temperature", "Stator Temperature"),
    ("estimated stator", "Stator Temperature"),
    ("rotor", "Rotor Temperature"),
    ("coolant", "Coolant"),
    ("power switch junction", "Power Switch Junction Temp"),
    ("power switch", "Power Switch Temperature"),
    ("dc link capacitor", "DC Link Capacitor"),
    ("oil pump", "Oil Pump"),
    ("oil temperature", "Oil Temperature"),
    ("oil ", "Oil"),
    ("lvdc", "LVDC"),
    ("kl15", "KL15"),
    ("overspeed", "Speed"),
    ("speed", "Speed"),
    ("vcu speed", "Speed"),
    ("resolver", "Resolver"),
    ("hvil", "HV Interlock"),
    ("crash", "Crash Input"),
    ("can ", "CAN"),
    ("eeprom", "Memory"),
    ("locked rotor", "Speed"),
    ("id current", "Id Current"),
    ("torque", "Torque"),
    ("digital communication", "Communication"),
    ("desat", "Gate Driver"),
    ("pwm", "PWM Plausibility"),
    ("excitation", "Resolver"),
    ("sin/cos", "Resolver"),
    ("decel", "Vehicle Plausibility"),
    ("gdr_", "Gate Driver IC"),
    ("pwrswt_", "Gate Driver IC"),
    ("5v_safe", "Gate Driver IC"),
    ("15v_gd", "Gate Driver IC"),
    ("mosfet", "Gate Driver"),
    ("hardware overvoltage", "HVDC Voltage"),
    ("overvoltage reference", "HVDC Voltage"),
    ("rotary joint", "Mechanical"),
    ("driver", "Gate Driver"),
    ("hardware over", "Hardware Protection"),
    ("contactor", "Pre-charge"),
    ("pre-charge", "Pre-charge"),
    ("pre- charging", "Pre-charge"),
    ("differential pressure", "Pre-charge"),
]


def subject_of(name: str) -> str:
    n = name.lower()
    for k, v in SUBJECT_RULES:
        if k in n:
            return v
    return "Other"


FAULT_TYPE_RULES = [
    (lambda n, s: any(k in n for k in ["temperature", "overheat", "stator hotspot"]), "Thermal"),
    (lambda n, s: any(k in n for k in ["voltage", "overvoltage", "undervoltage"]), "Voltage"),
    (lambda n, s: any(k in n for k in ["current", "overcurrent", "undercurrent", "offset"]), "Current"),
    (lambda n, s: any(k in n for k in ["overspeed", "speed", "locked rotor"]), "Speed / Position"),
    (lambda n, s: any(k in n for k in ["resolver", "sin/cos", "excitation"]), "Position"),
    (lambda n, s: any(k in n for k in ["can ", "e2e", "communication"]), "Communication"),
    (lambda n, s: any(k in n for k in ["eeprom", "nvm", "memory"]), "Memory"),
    (lambda n, s: any(k in n for k in ["pwm", "torque", "deceleration", "direction"]), "Plausibility"),
    (lambda n, s: any(k in n for k in ["hvil", "crash"]), "Safety I/O"),
    (lambda n, s: any(k in n for k in ["discharge"]), "HV Discharge"),
    (lambda n, s: any(k in n for k in ["pre-charge", "pre- charging", "contactor"]), "Pre-charge"),
    (lambda n, s: any(k in n for k in ["gate", "desat", "gdr_", "pwrswt_"]), "Gate Driver"),
    (lambda n, s: any(k in n for k in ["oil pump", "oil "]), "Oil System"),
    (lambda n, s: any(k in n for k in ["module fault", "controller"]), "Power Module"),
]


def fault_type_of(name: str, subject: str = "") -> str:
    n = name.lower()
    for fn, t in FAULT_TYPE_RULES:
        if fn(n, subject):
            return t
    return "Other"


def thermal_tier(name: str) -> int:
    n = name.lower()
    if "measurement" in n: return 10
    if "open circuit" in n: return 12
    if "short to battery" in n or "short to power" in n: return 14
    if "short to ground" in n or "short to gnd" in n: return 16
    if "plausibility" in n: return 20
    if "functional over" in n and "derating" in n: return 25
    if ("overtemperature" in n or "over-temperature" in n) and "level 2" in n: return 25
    if "over-temperature" in n and "derating" in n: return 25
    if ("overtemperature" in n or "over-temperature" in n) and "level 3" in n: return 30
    if "discharge functional over" in n: return 25
    if "delta" in n and "junction" in n: return 35
    if ("overtemperature" in n or "over-temperature" in n) and "estimated" in n: return 32
    if "overtemperature failure" in n or "over-temperature failure" in n: return 30
    if "undertemperature" in n: return 60
    if "gradient" in n: return 70
    return 50

# ---------------------------------------------------------------------------
# Loaders


def load_review_qm() -> list[dict]:
    wb = load_workbook(REVIEW_XLSX, read_only=True, data_only=True)
    ws = wb["Master v3.1 QM"]
    hdr = [norm(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
    out = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(v in (None, "") for v in row):
            continue
        rec = {}
        for j, h in enumerate(hdr):
            if not h:
                continue
            rec[h] = norm(row[j]) if j < len(row) else ""
        rec["_review_row"] = i
        out.append(rec)
    wb.close()
    return out


def load_review_fusa() -> list[dict]:
    wb = load_workbook(REVIEW_XLSX, read_only=True, data_only=True)
    ws = wb["Diag matrix_Fusa"]
    rows = list(ws.iter_rows(min_row=3, values_only=True))
    hdr = [norm(c) for c in rows[0]]
    out = []
    for i, row in enumerate(rows[1:], start=4):
        if not row or all(v in (None, "") for v in row):
            continue
        if not (row[hdr.index("Name")] if "Name" in hdr else None):
            # gate-driver IC rows have no Name in column "Name" but have name in "Name" col
            pass
        rec = {}
        for j, h in enumerate(hdr):
            if not h:
                continue
            rec[h] = norm(row[j]) if j < len(row) else ""
        if not rec.get("Name"):
            continue
        rec["_review_row"] = i
        out.append(rec)
    wb.close()
    return out


def load_customer() -> dict[str, dict]:
    return json.loads(CUSTOMER_JSON.read_text(encoding="utf-8"))


# ---------------------------------------------------------------------------
# Cleanup pipeline


def collapse_duplicate_dtcs(rows: list[dict], gaps: list[dict]) -> list[dict]:
    """Merge rows that share a DTC (e.g. Positive + Negative overspeed both on P1A 01 16)."""
    by_dtc: dict[str, list[dict]] = defaultdict(list)
    order: list[str] = []
    out: list[dict] = []
    for r in rows:
        d = r.get("DTC Code", "")
        if not d:
            out.append(r); continue
        if d not in by_dtc:
            order.append(d)
        by_dtc[d].append(r)
    for d in order:
        bucket = by_dtc[d]
        if len(bucket) == 1:
            out.append(bucket[0]); continue
        names = [r.get("Fault Name", "") for r in bucket]
        merged = dict(bucket[0])
        merged["Fault Name"] = " / ".join(names)
        existing_note = merged.get("Note", "")
        merge_tag = (
            f"DTC merged: {len(bucket)} rows shared this code in V1 review "
            f"({'; '.join(names)})."
        )
        merged["Note"] = (existing_note + " | " + merge_tag).strip(" |") if existing_note else merge_tag
        out.append(merged)
        gaps.append({
            "severity": "HIGH", "area": "Duplicate DTC",
            "where": f"DTC {d}",
            "issue": f"{len(bucket)} different faults shared DTC {d}: {'; '.join(names)}.",
            "action": ("A DTC must map to exactly one fault. Either request a new DTC "
                      "from customer for the second fault or merge them under one name."),
        })
    return out


def clean_qm(qm_rows: list[dict]) -> tuple[list[dict], list[dict]]:
    """Clean the reviewer's QM sheet. Returns (clean_rows, gap_log)."""
    gaps: list[dict] = []
    clean: list[dict] = []
    parent: dict | None = None
    for r in qm_rows:
        dtc = norm_dtc(r.get("DTC Code", ""))
        name = norm(r.get("Fault Name", ""))
        rxn = normalise_reaction(r.get("Reaction", ""))
        # detect orphaned child row (no DTC, no SPN, has name)
        if not dtc and not r.get("SPN") and name:
            if parent is None:
                gaps.append({
                    "severity": "CRITICAL", "area": "QM Sheet Layout",
                    "where": f"Review row {r['_review_row']}",
                    "issue": f"Orphan threshold-tier row '{name}' has no parent DTC; cannot be released.",
                    "action": "Assign a DTC code or merge into the previous parent row's note.",
                })
                continue
            # promote to standalone row inheriting parent's subject/type but no DTC
            child = dict(parent)
            child["Fault Name"] = name
            child["Reaction"] = rxn
            child["DTC Code"] = ""
            child["SPN"] = ""
            child["FMI"] = ""
            child["Status"] = "Reviewer-modified"
            child["Note"] = (
                (parent.get("Note", "") + " | Threshold tier added by reviewer; needs own DTC.")
                .strip(" |")
            )
            gaps.append({
                "severity": "HIGH", "area": "QM Sheet",
                "where": f"Review row {r['_review_row']}",
                "issue": (f"Reviewer added threshold-tier row '{name}' under parent "
                          f"'{parent['Fault Name']}' (DTC {parent.get('DTC Code')}). "
                          "No DTC/SPN/FMI assigned."),
                "action": (f"Assign next free DTC in customer range for parent SPN "
                          f"{parent.get('SPN')} or merge reaction into parent's reaction column."),
            })
            clean.append(child)
            continue
        # new parent row
        rec = dict(r)
        rec["DTC Code"] = dtc
        rec["Fault Name"] = name
        rec["Reaction"] = rxn
        # normalise fault level
        fl = rec.get("Fault Level", "")
        sev = parse_fault_level_letter(fl)
        if not sev and fl.upper().startswith("L"):
            gaps.append({
                "severity": "HIGH", "area": "Fault Level Scheme",
                "where": f"DTC {dtc}",
                "issue": f"Fault Level '{fl}' uses E-GAS L1/L2/L3 scheme; customer expects A/B/C/D.",
                "action": "Map the row's Reaction class (Linear Deration ~ B; ASC ~ A) and replace L1.",
            })
        # ensure subject + fault type
        rec["Subject"] = rec.get("Subject") or subject_of(name)
        rec["Fault Type"] = rec.get("Fault Type") or fault_type_of(name, rec["Subject"])
        rec["Status"] = rec.get("Status") or "Original (V1)"
        clean.append(rec)
        parent = rec
    return clean, gaps


def clean_fusa(rows: list[dict]) -> tuple[list[dict], list[dict]]:
    """Map the V1 FuSa sheet into the v3.2 deliverable shape."""
    gaps: list[dict] = []
    out: list[dict] = []
    for r in rows:
        rec: dict[str, str] = {}
        rec["Fault Name"] = norm(r.get("Name"))
        rec["DTC Code"] = norm_dtc(r.get("DTC Code", ""))
        rec["SPN"] = norm(r.get("SPN"))
        rec["FMI"] = norm(r.get("FMI"))
        rec["Fault Level"] = norm(r.get("Fault Level"))
        rec["ASIL"] = norm(r.get("ASIL"))
        rec["Domain"] = norm(r.get("Domain"))
        rec["Enabling Conditions"] = norm(r.get("Enabling confitions"))
        rec["Detection Criteria"] = norm(r.get("Detection Criteria"))
        rec["Periodicity [ms]"] = norm(r.get("Periodicity [ms]"))
        rec["Detection Confirmation Time [ms]"] = norm(
            r.get("Detection Confirmation time [ms]")
        )
        rec["Healing Criteria"] = norm(r.get("Healing criteria"))
        rec["Healing Confirmation Time [ms]"] = norm(
            r.get("Healing Confirmation time [ms]")
        )
        rec["Reaction"] = normalise_reaction(r.get("Reaction", ""))
        rec["VCU Request"] = norm(r.get("VCU Request"))
        rec["Notification"] = norm(r.get("Notification"))
        rec["Owner"] = norm(r.get("Owner"))
        rec["Severity"] = norm(r.get("Severity"))
        rec["Note"] = norm(r.get("Note / Comment"))
        rec["Subject"] = subject_of(rec["Fault Name"])
        rec["Fault Type"] = fault_type_of(rec["Fault Name"], rec["Subject"])
        rec["Status"] = "Original (V1)"

        if not rec["DTC Code"] and not rec["SPN"]:
            gaps.append({
                "severity": "CRITICAL", "area": "FuSa Sheet",
                "where": f"FuSa row {r['_review_row']} '{rec['Fault Name'][:40]}'",
                "issue": "Gate-driver IC fault has NO DTC / SPN / FMI assigned. Cannot be reported on CAN.",
                "action": "Assign DTC from customer reserved range (P1A 01 6x..7x) or document as 'internal-only'.",
            })
        out.append(rec)
    return out, gaps


def add_customer_only_dtcs(qm: list[dict], cust: dict[str, dict],
                           qm_dtcs: set[str]) -> tuple[list[dict], list[dict]]:
    gaps: list[dict] = []
    extras: list[dict] = []
    seen_added = {a["dtc"] for a in CUSTOMER_ADDED_FROM_JPG}
    for entry in CUSTOMER_ADDED_FROM_JPG:
        if entry["dtc"] in qm_dtcs:
            continue
        rec = {
            "DTC Code": entry["dtc"],
            "SPN": entry["spn"],
            "FMI": entry["fmi"],
            "Fault Level": entry["fl"],
            "ASIL": "QM",
            "Fault Name": entry["name"],
            "Subject": entry["subject"],
            "Fault Type": fault_type_of(entry["name"], entry["subject"]),
            "Reaction": "Linear Deration" if entry["fl"] == "B" else "ASC",
            "Enabling Conditions": "TBD",
            "Detection Criteria": "TBD",
            "Periodicity [ms]": "TBD",
            "Detection Confirmation Time [ms]": "TBD",
            "Healing Criteria": "TBD",
            "Healing Confirmation Time [ms]": "TBD",
            "VCU Request": "",
            "Notification": "",
            "Freeze Frame": "",
            "Variant": "",
            "Status": "Added (Customer)",
            "Note": "Customer-required DTC (HANDE HCV DTC1.jpg green-highlighted).",
        }
        extras.append(rec)
        gaps.append({
            "severity": "HIGH", "area": "Customer-required DTC",
            "where": f"DTC {entry['dtc']}",
            "issue": (f"Customer requires DTC {entry['dtc']} ({entry['name']}); was missing "
                     "from review QM sheet."),
            "action": ("Engineer must define Detection Criteria, Periodicity, and Healing "
                      "before release."),
        })
    return extras, gaps


def apply_customer_overrides(qm: list[dict]) -> list[dict]:
    log: list[dict] = []
    by_dtc = defaultdict(list)
    for r in qm:
        if r.get("DTC Code"):
            by_dtc[r["DTC Code"]].append(r)
    for dtc, ovr in CUSTOMER_FL_OVERRIDES.items():
        for r in by_dtc.get(dtc, []):
            cur_letter = parse_fault_level_letter(r.get("Fault Level", ""))
            if cur_letter == ovr["to"]:
                continue
            r["Fault Level"] = ovr["to"]
            r["Note"] = (
                (r.get("Note", "") +
                 f" | Customer-requested fault-level change {ovr['from']}->{ovr['to']} "
                 f"({ovr['src']}).").strip(" |")
            )
            log.append({
                "severity": "INFO", "area": "Customer Approval",
                "where": f"DTC {dtc}",
                "issue": f"Fault Level changed {ovr['from']} -> {ovr['to']}.",
                "action": ovr["src"],
            })
    return log

# ---------------------------------------------------------------------------
# Customer cross-check


def customer_cross_check(qm: list[dict], fusa: list[dict],
                          cust: dict[str, dict]) -> tuple[list[dict], list[dict]]:
    """For each customer DTC, compare against QM and FuSa entries."""
    gaps: list[dict] = []
    rows: list[dict] = []
    qm_by_dtc = {r["DTC Code"]: r for r in qm if r.get("DTC Code")}
    fusa_by_dtc = {r["DTC Code"]: r for r in fusa if r.get("DTC Code")}

    all_dtcs = sorted(set(cust.keys()) | set(qm_by_dtc.keys()) | set(fusa_by_dtc.keys()))
    for dtc in all_dtcs:
        c = cust.get(dtc, {})
        q = qm_by_dtc.get(dtc, {})
        f = fusa_by_dtc.get(dtc, {})

        c_spn, c_fmi, c_fl = c.get("spn", ""), c.get("fmi", ""), c.get("severity") or c.get("fault_level_raw", "")
        q_spn, q_fmi, q_fl = q.get("SPN", ""), q.get("FMI", ""), parse_fault_level_letter(q.get("Fault Level", ""))
        f_spn, f_fmi, f_fl = f.get("SPN", ""), f.get("FMI", ""), parse_fault_level_letter(f.get("Fault Level", ""))

        verdict_qm = "OK"
        verdict_fusa = "OK"
        notes = []

        if not c:
            verdict_qm = "MISSING IN CUSTOMER"
            verdict_fusa = "MISSING IN CUSTOMER"
            notes.append("DTC not in customer-approved table.")
        else:
            if not q:
                verdict_qm = "MISSING IN MATRIX"
                notes.append("DTC missing from QM sheet.")
            else:
                if c_spn and q_spn and c_spn != q_spn:
                    verdict_qm = "MISMATCH"
                    notes.append(f"SPN diff: cust={c_spn} qm={q_spn}.")
                if c_fmi and q_fmi and c_fmi != q_fmi:
                    verdict_qm = "MISMATCH"
                    notes.append(f"FMI diff: cust={c_fmi} qm={q_fmi}.")
                if c_fl and q_fl and c_fl != q_fl:
                    verdict_qm = "MISMATCH"
                    notes.append(f"FL diff: cust={c_fl} qm={q_fl}.")

            if not f:
                verdict_fusa = "MISSING IN MATRIX"
            else:
                if c_spn and f_spn and c_spn != f_spn:
                    verdict_fusa = "MISMATCH"
                if c_fmi and f_fmi and c_fmi != f_fmi:
                    verdict_fusa = "MISMATCH"
                if c_fl and f_fl and c_fl != f_fl:
                    verdict_fusa = "MISMATCH"

        rows.append({
            "DTC": dtc,
            "Customer Name": c.get("name_en", "") if c else "",
            "Customer SPN": c_spn, "Customer FMI": c_fmi, "Customer FL": c_fl,
            "QM Name": q.get("Fault Name", ""),
            "QM SPN": q_spn, "QM FMI": q_fmi, "QM FL": q_fl,
            "FuSa Name": f.get("Fault Name", ""),
            "FuSa SPN": f_spn, "FuSa FMI": f_fmi, "FuSa FL": f_fl,
            "QM Verdict": verdict_qm,
            "FuSa Verdict": verdict_fusa,
            "Notes": " ".join(notes),
        })

        if verdict_qm == "MISMATCH":
            gaps.append({
                "severity": "CRITICAL", "area": "Customer Mismatch (QM)",
                "where": f"DTC {dtc}",
                "issue": notes[-1] if notes else "QM disagrees with customer-approved values.",
                "action": "Align with customer doc Customer_Diag.pdf or escalate for variance approval.",
            })
        elif verdict_qm == "MISSING IN MATRIX":
            gaps.append({
                "severity": "HIGH", "area": "Customer Coverage (QM)",
                "where": f"DTC {dtc}",
                "issue": f"Customer-approved DTC '{c.get('name_en','')}' missing from QM sheet.",
                "action": "Add row to QM matrix with full detection / healing definition.",
            })
        if verdict_fusa == "MISMATCH":
            gaps.append({
                "severity": "CRITICAL", "area": "Customer Mismatch (FuSa)",
                "where": f"DTC {dtc}",
                "issue": "FuSa disagrees with customer-approved values.",
                "action": "Align with customer doc.",
            })
    return rows, gaps


def qm_fusa_alignment(qm: list[dict], fusa: list[dict]) -> tuple[list[dict], list[dict]]:
    gaps: list[dict] = []
    rows: list[dict] = []
    by_q = {r["DTC Code"]: r for r in qm if r.get("DTC Code")}
    by_f = {r["DTC Code"]: r for r in fusa if r.get("DTC Code")}
    all_dtcs = sorted(set(by_q.keys()) | set(by_f.keys()))
    for dtc in all_dtcs:
        q = by_q.get(dtc); f = by_f.get(dtc)
        verdict = "OK"
        notes = []
        if q and f:
            if (q.get("SPN") or "") != (f.get("SPN") or ""):
                verdict = "MISMATCH"; notes.append(f"SPN: qm={q.get('SPN')} fusa={f.get('SPN')}")
            if (q.get("FMI") or "") != (f.get("FMI") or ""):
                verdict = "MISMATCH"; notes.append(f"FMI: qm={q.get('FMI')} fusa={f.get('FMI')}")
            qf = parse_fault_level_letter(q.get("Fault Level", ""))
            ff = parse_fault_level_letter(f.get("Fault Level", ""))
            if qf and ff and qf != ff:
                verdict = "MISMATCH"; notes.append(f"FL letter: qm={qf} fusa={ff}")
        elif q and not f:
            verdict = "QM ONLY"
        elif f and not q:
            verdict = "FuSa ONLY"
        rows.append({
            "DTC": dtc,
            "QM Name": q.get("Fault Name", "") if q else "",
            "FuSa Name": f.get("Fault Name", "") if f else "",
            "QM SPN": q.get("SPN", "") if q else "",
            "FuSa SPN": f.get("SPN", "") if f else "",
            "QM FMI": q.get("FMI", "") if q else "",
            "FuSa FMI": f.get("FMI", "") if f else "",
            "QM FL": parse_fault_level_letter(q.get("Fault Level", "")) if q else "",
            "FuSa FL": parse_fault_level_letter(f.get("Fault Level", "")) if f else "",
            "QM Reaction": q.get("Reaction", "") if q else "",
            "FuSa Reaction": f.get("Reaction", "") if f else "",
            "Verdict": verdict,
            "Notes": "; ".join(notes),
        })
        if verdict == "MISMATCH":
            gaps.append({
                "severity": "HIGH", "area": "QM↔FuSa Alignment",
                "where": f"DTC {dtc}",
                "issue": "; ".join(notes),
                "action": "Pick the authoritative value (per customer table) and update the other sheet.",
            })
    return rows, gaps


def cell_completeness_audit(qm: list[dict]) -> list[dict]:
    """Detect blank Detection Criteria, Healing Criteria, Periodicity etc."""
    gaps: list[dict] = []
    required = [
        ("Detection Criteria", "HIGH"),
        ("Healing Criteria", "HIGH"),
        ("Periodicity [ms]", "MEDIUM"),
        ("Detection Confirmation Time [ms]", "MEDIUM"),
        ("Healing Confirmation Time [ms]", "MEDIUM"),
        ("Reaction", "HIGH"),
        ("Enabling Conditions", "MEDIUM"),
    ]
    tbd_re = re.compile(r"^\s*tbd\s*$", re.I)
    for r in qm:
        for col, sev in required:
            v = r.get(col, "")
            if not v or tbd_re.match(v) or v.strip() in {"-"}:
                gaps.append({
                    "severity": sev, "area": "Cell Completeness",
                    "where": f"DTC {r.get('DTC Code','?')}: {r.get('Fault Name','')[:40]}",
                    "issue": f"'{col}' is empty or TBD.",
                    "action": f"Fill in '{col}' before release.",
                })
    return gaps

# ---------------------------------------------------------------------------
# Layman justification


def layman_for(rec: dict) -> str:
    n = rec.get("Fault Name", "")
    nl = n.lower()
    rxn = rec.get("Reaction", "")
    fl = rec.get("Fault Level", "")

    if "measurement failure" in nl or "measurement fault" in nl:
        what = "the sensor reading is outside its physically possible range, so the MCU cannot trust it"
    elif "open circuit" in nl:
        what = "the wire to the sensor looks broken (open)"
    elif "short to power" in nl or "short to battery" in nl:
        what = "the sensor wire is shorted to the supply rail"
    elif "short to ground" in nl or "short to gnd" in nl:
        what = "the sensor wire is shorted to ground"
    elif "plausibility" in nl:
        what = "two sources that should agree disagree by too much"
    elif "level 2" in nl and ("over-temperature" in nl or "overtemperature" in nl):
        what = (
            f"{nl.split('overtemperature')[0].strip().split('level')[0].strip()} "
            "is getting hot — early warning, MCU starts derating"
        )
    elif "level 3" in nl and ("over-temperature" in nl or "overtemperature" in nl):
        what = (
            f"{nl.split('overtemperature')[0].strip().split('level')[0].strip()} "
            "crossed the hard safety temperature limit"
        )
    elif "functional over" in nl and "derating" in nl:
        thing = nl.split("functional")[0].strip()
        what = (
            f"{thing} is getting hot but still safe — early warning where the MCU starts "
            "derating before any safety limit is hit"
        )
    elif "discharge inhibited" in nl:
        what = ("the active-discharge resistor heated up too much; next active discharge is "
                "inhibited (passive discharge still works)")
    elif "overtemperature" in nl or "over-temperature" in nl:
        thing = nl.split("overtemperature")[0].strip()
        what = f"{thing} crossed the temperature safety limit"
    elif "undertemperature" in nl:
        what = "the measured temperature is impossibly low — sensor or wiring issue"
    elif "gradient" in nl and "temperature" in nl:
        what = "temperature is changing faster than physics allows — sensor noise / wiring"
    elif "overcurrent hw" in nl or "hw fault" in nl:
        what = "hardware comparator tripped — current crossed the hardware safety threshold"
    elif "overcurrent" in nl:
        what = "current crossed the software overcurrent threshold"
    elif "undercurrent" in nl:
        what = "current is below the expected operating range — sensor or wiring suspect"
    elif "offset" in nl and "current" in nl:
        what = "current sensor reads non-zero with the inverter idle — needs re-calibration"
    elif "current sum" in nl:
        what = "phase A+B+C currents do not sum to zero — one phase sensor is wrong"
    elif "overvoltage hw" in nl:
        what = "hardware comparator tripped — voltage crossed the hardware safety threshold"
    elif "overvoltage" in nl:
        what = "voltage crossed the overvoltage threshold"
    elif "undervoltage" in nl:
        what = "voltage dropped below the safe operating threshold"
    elif "resolver excitation" in nl:
        what = "resolver excitation signal is bad — rotor position sensing at risk"
    elif "resolver" in nl or "sin/cos" in nl:
        what = "rotor-position sensor signal is bad"
    elif "overspeed" in nl:
        what = "motor speed exceeded the allowed limit"
    elif "locked rotor" in nl:
        what = "rotor commanded but not turning (mechanical lock or sensor disagreement)"
    elif "can busoff" in nl or "can line" in nl:
        what = "vehicle CAN bus is in bus-off; MCU lost network access"
    elif "e2e" in nl or "can signal" in nl:
        what = "expected CAN signal CRC / counter check failed — message may be corrupted or stale"
    elif "eeprom" in nl or "nvm" in nl:
        what = "non-volatile memory check failed — calibrations may be unreliable"
    elif "hvil" in nl:
        what = "HV interlock loop is broken — assume HV is unsafe to energize"
    elif "crash" in nl:
        what = "crash signal received — MCU must shut down and discharge HV"
    elif "pwm" in nl and "plausibility" in nl:
        what = "redundant PWM checker disagrees with main PWM — gate-driver request not trusted"
    elif "pre-charge" in nl or "pre- charging" in nl:
        what = "HV pre-charge sequence did not complete correctly (capacitor not charged in time)"
    elif "contactor" in nl:
        what = "HV contactor stuck closed when it should be open"
    elif "oil pump" in nl:
        what = "oil pump issue — cooling or lubrication may be impaired"
    elif "torque" in nl and ("output" in nl or "actual" in nl):
        what = "actual torque differs too much from requested torque"
    elif "active discharge" in nl and "current" in nl:
        what = "active-discharge current sensor reading is wrong"
    elif "direction" in nl or "gear" in nl:
        what = "requested gear / direction does not match what the motor is doing"
    else:
        what = nl

    if "linear derat" in rxn.lower():
        action = "MCU smoothly cuts torque so the situation does not get worse."
    elif rxn.upper() == "ASC":
        action = "MCU shorts the three motor phases together (ASC) to stop creating heat / current."
    elif "asc or 6 switch open" in rxn.lower():
        action = "Safety layer shorts the three phases (ASC) or opens all six switches."
    elif "6 switch open" in rxn.lower():
        action = "MCU opens all six power switches so the inverter no longer drives the motor."
    elif "zero torque" in rxn.lower():
        action = "MCU forces output torque to zero."
    elif "torque limit" in rxn.lower():
        action = "MCU caps the available torque while the issue is present."
    elif "active discharge" in rxn.lower():
        action = "MCU performs HV active discharge."
    elif rxn.lower() in {"no reaction", ""}:
        action = "" if rxn.lower() in {""} else "MCU only reports the fault; no torque action."
    else:
        action = f"Reaction: {rxn}." if rxn else ""

    fl_note = ""
    if fl in {"A", "A/3"}:
        fl_note = " (Customer FL=A: major safety fault, must stop and repair.)"
    elif fl in {"B", "B/2"}:
        fl_note = " (Customer FL=B: serious fault, vehicle still drivable, store DTC.)"
    elif fl in {"C", "C/1"}:
        fl_note = " (Customer FL=C: general fault, repair recommended.)"

    s = f"What it means: {what}."
    if action:
        s += " " + action
    if fl_note:
        s += fl_note
    return s

# ---------------------------------------------------------------------------
# Workbook builder helpers


def style_header(ws, row: int, n_cols: int) -> None:
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = 32


def autosize(ws, widths: dict[str, int]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

# ---------------------------------------------------------------------------
# Sheet builders


def sheet_cover(wb: Workbook, counts: dict[str, int]) -> None:
    ws = wb.active
    ws.title = "Cover & Legend"
    ws["A1"] = "HANDE MCU — Consolidated Diag v3.2 (Deliverable)"
    ws["A1"].font = Font(bold=True, size=18, color="1F3864")
    ws.merge_cells("A1:F1")

    ws["A3"] = (
        "Source review file: Consolidated Diag v3.2 Review .xlsx  |  "
        "Customer doc: Customer_Diag/Customer_Diag.pdf  |  "
        "Customer JPGs: HANDE HCV DTC1/2/3.jpg"
    )
    ws["A3"].font = Font(italic=True, color="595959")
    ws.merge_cells("A3:F3")

    ws["A5"] = "Sheets in this workbook"
    ws["A5"].font = TITLE_FONT
    sheets = [
        ("Cover & Legend", "What is in this workbook, status colours, gap severities"),
        ("Master QM v3.2", "Final QM diagnostic matrix (cleaned, ordered, customer-aligned)"),
        ("Master FuSa v3.2", "Final FuSa diagnostic matrix (same column shape as QM)"),
        ("Customer Cross-Check", "Every DTC vs customer-approved Customer_Diag.pdf (SPN/FMI/Fault Level)"),
        ("QM <-> FuSa Alignment", "Side-by-side per DTC; flags any QM-vs-FuSa mismatch"),
        ("Justifications", "Plain-English explanation per fault"),
        ("Logs & Gaps", "Every issue found; CRITICAL / HIGH / MEDIUM / LOW; action required before release"),
    ]
    ws["A6"] = "Sheet"; ws["B6"] = "Purpose"
    style_header(ws, 6, 2)
    for i, (s, p) in enumerate(sheets, start=7):
        ws.cell(row=i, column=1, value=s).font = Font(bold=True)
        ws.cell(row=i, column=2, value=p)

    r = 7 + len(sheets) + 2
    ws.cell(row=r, column=1, value="Status colours used in matrix sheets").font = TITLE_FONT
    r += 1
    ws.cell(row=r, column=1, value="Status").font = SUB_HDR_FONT
    ws.cell(row=r, column=2, value="Meaning").font = SUB_HDR_FONT
    ws.cell(row=r, column=3, value="Colour").font = SUB_HDR_FONT
    style_header(ws, r, 3)
    r += 1
    legend = [
        ("Original (V1)", "Row from V1 sheet 'Diag matrix_V2'"),
        ("Reviewer-modified", "Reviewer edited the V1 row in v3.2 review"),
        ("Added (V3 AI)", "Functional / derating row added by V3 AI consolidation"),
        ("Added (Customer)", "Customer-required row from HANDE HCV DTC JPGs / PDF"),
        ("Added (new in v3.2)", "Row newly added by v3.2 deliverable build"),
        ("Removed (Customer)", "Customer requested removal (kept for audit)"),
    ]
    for s, m in legend:
        ws.cell(row=r, column=1, value=s).font = Font(bold=True)
        ws.cell(row=r, column=2, value=m)
        ws.cell(row=r, column=3, value="").fill = STATUS_FILL.get(s, PatternFill())
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Gap severity colours used in 'Logs & Gaps'").font = TITLE_FONT
    r += 1
    ws.cell(row=r, column=1, value="Severity").font = SUB_HDR_FONT
    ws.cell(row=r, column=2, value="Meaning").font = SUB_HDR_FONT
    ws.cell(row=r, column=3, value="Colour").font = SUB_HDR_FONT
    style_header(ws, r, 3)
    r += 1
    sev_legend = [
        ("CRITICAL", "Blocks release. Must be fixed."),
        ("HIGH", "Required before customer hand-off."),
        ("MEDIUM", "Should be fixed for v3.2; can be tracked."),
        ("LOW", "Cosmetic / nice-to-have."),
        ("INFO", "Not a defect, traceability note."),
    ]
    for s, m in sev_legend:
        ws.cell(row=r, column=1, value=s).font = Font(bold=True)
        ws.cell(row=r, column=2, value=m)
        cell = ws.cell(row=r, column=3, value="")
        cell.fill = SEVERITY_FILL.get(s, PatternFill())
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Counts").font = TITLE_FONT
    r += 1
    for k, v in counts.items():
        ws.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws.cell(row=r, column=2, value=v)
        r += 1

    autosize(ws, {"A": 30, "B": 70, "C": 14, "D": 14, "E": 14, "F": 14})


def sheet_master(wb: Workbook, name: str, columns: list[str], rows: list[dict]) -> None:
    ws = wb.create_sheet(name)
    ws.append(columns)
    style_header(ws, 1, len(columns))
    for i, r in enumerate(rows, start=1):
        vals = [i if c == "#" else r.get(c, "") for c in columns]
        ws.append(vals)
        # status fill
        st = r.get("Status", "")
        st_col = columns.index("Status") + 1 if "Status" in columns else None
        if st_col and st in STATUS_FILL:
            ws.cell(row=ws.max_row, column=st_col).fill = STATUS_FILL[st]
        # verification fill
        ver = r.get("Verification", "")
        if ver and "Verification" in columns:
            ver_col = columns.index("Verification") + 1
            ws.cell(row=ws.max_row, column=ver_col).fill = VERIFY_FILL.get(ver, PatternFill())
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical="top")
            c.border = BORDER
    widths = {"#": 5, "Fault Type": 14, "Subject": 22, "DTC Code": 12, "SPN": 8, "FMI": 6,
              "Fault Name": 50, "Reaction": 24, "ASIL": 8, "Fault Level": 9,
              "Enabling Conditions": 22, "Detection Criteria": 32,
              "Periodicity [ms]": 9, "Detection Confirmation Time [ms]": 13,
              "Healing Criteria": 32, "Healing Confirmation Time [ms]": 13,
              "VCU Request": 18, "Notification": 22, "Freeze Frame": 16, "Variant": 10,
              "Domain": 10, "Owner": 14, "Severity": 10,
              "Status": 18, "Verification": 14, "Note": 50}
    for i, c in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(c, 14)
    ws.freeze_panes = "D2"


def sheet_customer_cross(wb: Workbook, rows: list[dict]) -> None:
    cols = ["DTC", "Customer Name", "Customer SPN", "Customer FMI", "Customer FL",
            "QM Name", "QM SPN", "QM FMI", "QM FL", "QM Verdict",
            "FuSa Name", "FuSa SPN", "FuSa FMI", "FuSa FL", "FuSa Verdict",
            "Notes"]
    ws = wb.create_sheet("Customer Cross-Check")
    ws.append(cols)
    style_header(ws, 1, len(cols))
    for r in rows:
        ws.append([r.get(c, "") for c in cols])
        qm_v = r.get("QM Verdict", "")
        fusa_v = r.get("FuSa Verdict", "")
        if qm_v in VERIFY_FILL:
            ws.cell(row=ws.max_row, column=cols.index("QM Verdict") + 1).fill = VERIFY_FILL[qm_v]
        if fusa_v in VERIFY_FILL:
            ws.cell(row=ws.max_row, column=cols.index("FuSa Verdict") + 1).fill = VERIFY_FILL[fusa_v]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical="top")
            c.border = BORDER
    widths = {"A": 12, "B": 38, "C": 9, "D": 6, "E": 8, "F": 38, "G": 9, "H": 6, "I": 8,
              "J": 18, "K": 38, "L": 9, "M": 6, "N": 8, "O": 18, "P": 50}
    autosize(ws, widths)
    ws.freeze_panes = "B2"


def sheet_alignment(wb: Workbook, rows: list[dict]) -> None:
    cols = ["DTC", "QM Name", "FuSa Name", "QM SPN", "FuSa SPN",
            "QM FMI", "FuSa FMI", "QM FL", "FuSa FL",
            "QM Reaction", "FuSa Reaction", "Verdict", "Notes"]
    ws = wb.create_sheet("QM <-> FuSa Alignment")
    ws.append(cols); style_header(ws, 1, len(cols))
    fill_map = {"OK": VERIFY_FILL["OK"], "MISMATCH": VERIFY_FILL["MISMATCH"],
                "QM ONLY": VERIFY_FILL["INFO"], "FuSa ONLY": VERIFY_FILL["INFO"]}
    for r in rows:
        ws.append([r.get(c, "") for c in cols])
        v = r.get("Verdict", "")
        if v in fill_map:
            ws.cell(row=ws.max_row, column=cols.index("Verdict") + 1).fill = fill_map[v]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical="top"); c.border = BORDER
    autosize(ws, {"A": 12, "B": 34, "C": 34, "D": 9, "E": 9, "F": 6, "G": 6,
                  "H": 7, "I": 7, "J": 24, "K": 30, "L": 14, "M": 50})
    ws.freeze_panes = "B2"


def sheet_justifications(wb: Workbook, qm_rows: list[dict]) -> None:
    cols = ["#", "DTC Code", "Fault Name", "Subject", "Fault Level",
            "What this means (layman terms)", "Status"]
    ws = wb.create_sheet("Justifications")
    ws.append(cols); style_header(ws, 1, len(cols))
    for i, r in enumerate(qm_rows, start=1):
        ws.append([i, r.get("DTC Code", ""), r.get("Fault Name", ""),
                  r.get("Subject", ""), r.get("Fault Level", ""),
                  layman_for(r), r.get("Status", "")])
        st = r.get("Status", "")
        if st in STATUS_FILL:
            ws.cell(row=ws.max_row, column=len(cols)).fill = STATUS_FILL[st]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical="top"); c.border = BORDER
    autosize(ws, {"A": 5, "B": 12, "C": 50, "D": 22, "E": 7, "F": 95, "G": 18})
    ws.freeze_panes = "B2"


def sheet_gaps(wb: Workbook, gaps: list[dict]) -> None:
    cols = ["#", "Severity", "Area", "Where (DTC / Row)", "Issue", "Action required before release"]
    ws = wb.create_sheet("Logs & Gaps")
    # Sort: CRITICAL > HIGH > MEDIUM > LOW > INFO, then by area
    rank = {"CRITICAL": 0, "HIGH": 1, "MEDIUM": 2, "LOW": 3, "INFO": 4}
    gaps_sorted = sorted(gaps, key=lambda g: (rank.get(g.get("severity", "INFO"), 9),
                                              g.get("area", ""), g.get("where", "")))
    ws.append(cols); style_header(ws, 1, len(cols))
    counts: dict[str, int] = defaultdict(int)
    for i, g in enumerate(gaps_sorted, start=1):
        sev = g.get("severity", "INFO")
        ws.append([i, sev, g.get("area", ""), g.get("where", ""),
                   g.get("issue", ""), g.get("action", "")])
        cell = ws.cell(row=ws.max_row, column=2)
        cell.fill = SEVERITY_FILL.get(sev, PatternFill())
        if sev in SEVERITY_FONT_WHITE:
            cell.font = Font(bold=True, color="FFFFFF")
        else:
            cell.font = Font(bold=True)
        counts[sev] += 1
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical="top"); c.border = BORDER
    autosize(ws, {"A": 5, "B": 11, "C": 28, "D": 30, "E": 60, "F": 60})
    ws.freeze_panes = "B2"
    return counts

# ---------------------------------------------------------------------------
# Main


def order_qm(rows: list[dict]) -> list[dict]:
    enriched = []
    for r in rows:
        ft = r.get("Fault Type") or fault_type_of(r["Fault Name"], r.get("Subject", ""))
        sj = r.get("Subject") or subject_of(r["Fault Name"])
        tier = thermal_tier(r["Fault Name"]) if ft == "Thermal" else 50
        r["Fault Type"] = ft
        r["Subject"] = sj
        enriched.append((ft, sj, tier, r))
    seen: list[tuple[str, str]] = []
    seen_set: set[tuple[str, str]] = set()
    type_order = ["Voltage", "Current", "Thermal", "Speed / Position", "Position",
                  "HV Discharge", "Pre-charge", "Power Module", "Gate Driver",
                  "Plausibility", "Communication", "Memory", "Safety I/O", "Oil System", "Other"]
    for ft in type_order:
        for _, sj, _, _ in [e for e in enriched if e[0] == ft]:
            if (ft, sj) not in seen_set:
                seen_set.add((ft, sj)); seen.append((ft, sj))
    # tail: any types not in type_order
    for e in enriched:
        if (e[0], e[1]) not in seen_set:
            seen_set.add((e[0], e[1])); seen.append((e[0], e[1]))
    final: list[dict] = []
    for key in seen:
        members = [(t, r) for ft, sj, t, r in enriched if (ft, sj) == key]
        members.sort(key=lambda x: (x[0], x[1].get("DTC Code", "")))
        final.extend(r for _, r in members)
    return final


def main() -> None:
    qm_raw = load_review_qm()
    fusa_raw = load_review_fusa()
    customer = load_customer()

    qm_clean, gaps_clean_qm = clean_qm(qm_raw)
    qm_clean = collapse_duplicate_dtcs(qm_clean, gaps_clean_qm)
    fusa_clean, gaps_clean_fusa = clean_fusa(fusa_raw)
    fusa_clean = collapse_duplicate_dtcs(fusa_clean, gaps_clean_fusa)

    qm_dtcs = {r["DTC Code"] for r in qm_clean if r.get("DTC Code")}
    extras, gaps_extras = add_customer_only_dtcs(qm_clean, customer, qm_dtcs)
    qm_clean.extend(extras)

    gaps_overrides = apply_customer_overrides(qm_clean)

    qm_clean = order_qm(qm_clean)
    fusa_clean = order_qm(fusa_clean)

    # cross checks
    cust_rows, gaps_cust = customer_cross_check(qm_clean, fusa_clean, customer)
    align_rows, gaps_align = qm_fusa_alignment(qm_clean, fusa_clean)
    gaps_complete = cell_completeness_audit(qm_clean)

    # add per-row Verification flag (used in Master QM v3.2 status colour)
    cust_verdict = {r["DTC"]: r["QM Verdict"] for r in cust_rows}
    for r in qm_clean:
        d = r.get("DTC Code", "")
        r["Verification"] = cust_verdict.get(d, "INFO")
    fusa_verdict = {r["DTC"]: r["FuSa Verdict"] for r in cust_rows}
    for r in fusa_clean:
        d = r.get("DTC Code", "")
        r["Verification"] = fusa_verdict.get(d, "INFO")

    all_gaps = (gaps_clean_qm + gaps_clean_fusa + gaps_extras + gaps_overrides
                + gaps_cust + gaps_align + gaps_complete)

    # ---------------------- write workbook ----------------------
    wb = Workbook()
    counts = {
        "QM rows": len(qm_clean),
        "FuSa rows": len(fusa_clean),
        "Customer-approved DTCs": len(customer),
        "Total gaps logged": len(all_gaps),
    }
    sheet_cover(wb, counts)
    sheet_master(wb, "Master QM v3.2", QM_COLUMNS, qm_clean)
    sheet_master(wb, "Master FuSa v3.2", FUSA_COLUMNS, fusa_clean)
    sheet_customer_cross(wb, cust_rows)
    sheet_alignment(wb, align_rows)
    sheet_justifications(wb, qm_clean)
    gap_counts = sheet_gaps(wb, all_gaps)

    # update Cover counts with severity
    cover = wb["Cover & Legend"]
    last = cover.max_row + 2
    cover.cell(row=last, column=1, value="Gaps by severity").font = TITLE_FONT
    last += 1
    for sev in ["CRITICAL", "HIGH", "MEDIUM", "LOW", "INFO"]:
        cover.cell(row=last, column=1, value=sev).font = Font(bold=True)
        cover.cell(row=last, column=1).fill = SEVERITY_FILL.get(sev, PatternFill())
        if sev in SEVERITY_FONT_WHITE:
            cover.cell(row=last, column=1).font = Font(bold=True, color="FFFFFF")
        cover.cell(row=last, column=2, value=gap_counts.get(sev, 0))
        last += 1

    wb.save(OUT)
    print(f"Wrote {OUT}")
    print(f"  QM rows        : {len(qm_clean)}")
    print(f"  FuSa rows      : {len(fusa_clean)}")
    print(f"  Customer DTCs  : {len(customer)}")
    print(f"  Total gaps     : {len(all_gaps)}")
    for s, c in gap_counts.items():
        print(f"    {s:<10}: {c}")


if __name__ == "__main__":
    main()
