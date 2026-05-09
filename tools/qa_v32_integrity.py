"""Integrity & sanity checks on Consolidated Diag v3.2.xlsx."""
from __future__ import annotations
import io
import sys
from collections import Counter

import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

p = r"C:\Projects\Diag\Consolidated Diag v3.2.xlsx"
wb = openpyxl.load_workbook(p, read_only=True, data_only=True)

print("===== INTEGRITY CHECKS =====\n")

# 1) duplicate DTCs in QM (only count where DTC non-empty)
ws = wb["Master QM v3.2"]
hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
dtc_i = hdr.index("DTC Code"); name_i = hdr.index("Fault Name")
dtcs = []
for r in ws.iter_rows(min_row=2, values_only=True):
    if r[dtc_i]:
        dtcs.append(r[dtc_i])
dup = [k for k, v in Counter(dtcs).items() if v > 1]
print(f"[1] QM duplicate DTCs: {len(dup)}")
for d in dup:
    print(f"    -> {d}")

# 2) FuSa rows missing DTC
ws = wb["Master FuSa v3.2"]
hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
dtc_i = hdr.index("DTC Code"); name_i = hdr.index("Fault Name")
no_dtc = []
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r[dtc_i] and r[name_i]:
        no_dtc.append(r[name_i])
print(f"\n[2] FuSa rows without DTC: {len(no_dtc)}")
for n in no_dtc[:6]:
    print(f"    -> {n[:60]}")
if len(no_dtc) > 6:
    print(f"    ... and {len(no_dtc) - 6} more")

# 3) Customer Cross-Check verdict counts
ws = wb["Customer Cross-Check"]
hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
qm_v_i = hdr.index("QM Verdict"); fusa_v_i = hdr.index("FuSa Verdict")
qm_c, fusa_c = Counter(), Counter()
for r in ws.iter_rows(min_row=2, values_only=True):
    qm_c[r[qm_v_i]] += 1
    fusa_c[r[fusa_v_i]] += 1
print("\n[3] Customer Cross-Check verdict distribution:")
print("    QM   :", dict(qm_c))
print("    FuSa :", dict(fusa_c))

# 4) QM<->FuSa Alignment
ws = wb["QM <-> FuSa Alignment"]
hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
v_i = hdr.index("Verdict")
align_c = Counter()
for r in ws.iter_rows(min_row=2, values_only=True):
    align_c[r[v_i]] += 1
print("\n[4] QM<->FuSa Alignment distribution:", dict(align_c))

# 5) Logs & Gaps area distribution
ws = wb["Logs & Gaps"]
hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
sev_i = hdr.index("Severity"); area_i = hdr.index("Area")
ar_c = Counter()
for r in ws.iter_rows(min_row=2, values_only=True):
    ar_c[(r[sev_i], r[area_i])] += 1
print("\n[5] Gaps by (severity, area):")
for (s, a), c in sorted(ar_c.items(), key=lambda kv: (-kv[1], kv[0])):
    print(f"    {s:<10} {a:<32} {c}")

# 6) Stator block tier verification
ws = wb["Master QM v3.2"]
hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
sub_i = hdr.index("Subject"); name_i = hdr.index("Fault Name")
print("\n[6] Stator Temperature row sequence (must show derating BEFORE hard OT):")
seq = []
for r in ws.iter_rows(min_row=2, values_only=True):
    if r[sub_i] == "Stator Temperature":
        seq.append(str(r[name_i] or "")[:60])
for i, n in enumerate(seq, 1):
    print(f"    {i:2}. {n}")

# 7) Sanity: every Customer-added row should have FL ∈ {A,B,C}
ws = wb["Master QM v3.2"]
hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
st_i = hdr.index("Status"); fl_i = hdr.index("Fault Level"); name_i = hdr.index("Fault Name")
bad_fl = []
for r in ws.iter_rows(min_row=2, values_only=True):
    if r[st_i] == "Added (Customer)":
        if r[fl_i] not in ("A", "B", "C", "D"):
            bad_fl.append((r[name_i], r[fl_i]))
print(f"\n[7] Customer-added rows with non-A/B/C fault level: {len(bad_fl)}")
for n, f in bad_fl:
    print(f"    -> {n[:50]} | FL={f}")

wb.close()
print("\n===== INTEGRITY CHECKS COMPLETE =====")
