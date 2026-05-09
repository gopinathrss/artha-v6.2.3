import openpyxl

p = r"C:\Projects\Diag\Consolidated Diag v3.2.xlsx"
wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
print("SHEETS:")
for s in wb.sheetnames:
    ws = wb[s]
    print(f"  {s}: {ws.max_row} rows x {ws.max_column} cols")

ws = wb["Master QM v3.2"]
hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]


def gi(n: str) -> int:
    return hdr.index(n)


print("\n=== Stator Temperature block (QM) ===")
sub_i = gi("Subject"); name_i = gi("Fault Name"); dtc_i = gi("DTC Code")
rxn_i = gi("Reaction"); st_i = gi("Status"); fl_i = gi("Fault Level")
ver_i = gi("Verification"); spn_i = gi("SPN"); fmi_i = gi("FMI")
for r in ws.iter_rows(min_row=2, values_only=True):
    if r[sub_i] == "Stator Temperature":
        nm = str(r[name_i] or "")[:55]
        rxn = str(r[rxn_i] or "")[:25]
        print(f"  {str(r[dtc_i] or ''):<11} | SPN={str(r[spn_i] or ''):<7} FMI={str(r[fmi_i] or ''):<3} FL={str(r[fl_i] or ''):<3} | {nm:<55} | rxn={rxn:<25} | {str(r[st_i] or ''):<20} | {str(r[ver_i] or '')}")

print("\n=== HVDC Voltage block (QM) ===")
for r in ws.iter_rows(min_row=2, values_only=True):
    if r[sub_i] == "HVDC Voltage":
        nm = str(r[name_i] or "")[:55]
        rxn = str(r[rxn_i] or "")[:25]
        print(f"  {str(r[dtc_i] or ''):<11} | SPN={str(r[spn_i] or ''):<7} FMI={str(r[fmi_i] or ''):<3} FL={str(r[fl_i] or ''):<3} | {nm:<55} | rxn={rxn:<25} | {str(r[st_i] or ''):<20} | {str(r[ver_i] or '')}")

print("\n=== Top 12 CRITICAL gaps ===")
ws = wb["Logs & Gaps"]
n = 0
for r in ws.iter_rows(min_row=2, values_only=True):
    if (r[1] or "") == "CRITICAL":
        issue = str(r[4] or "")[:90]
        print(f"  #{r[0]:>3} | {str(r[2]):<28} | {str(r[3] or ''):<30} | {issue}")
        n += 1
        if n >= 12:
            break

print("\n=== Top 12 HIGH gaps ===")
n = 0
for r in ws.iter_rows(min_row=2, values_only=True):
    if (r[1] or "") == "HIGH":
        issue = str(r[4] or "")[:90]
        print(f"  #{r[0]:>3} | {str(r[2]):<28} | {str(r[3] or ''):<30} | {issue}")
        n += 1
        if n >= 12:
            break

print("\n=== Customer Cross-Check sample (first 12 mismatches) ===")
ws = wb["Customer Cross-Check"]
n = 0
for r in ws.iter_rows(min_row=2, values_only=True):
    if r[9] in ("MISMATCH", "MISSING IN MATRIX"):
        print(f"  {r[0]} | cust SPN={r[2]} FMI={r[3]} FL={r[4]} | qm SPN={r[6]} FMI={r[7]} FL={r[8]} | verdict={r[9]}")
        n += 1
        if n >= 12:
            break
wb.close()
