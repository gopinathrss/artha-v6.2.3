"""Build Consolidated Diag v3.1.xlsx.

Source of truth:
  - V1 sheet `Diag matrix_V2` in `Consolidated Diagnostic Excel_ v1 (1).xlsx`
    (~222 rows, contains both QM and FuSa flavours; deduplicated here)
V3 CSV is used only as a reference for SPN/FMI/DTC corrections and to pull in
net-new V3-AI rows that do not exist in V1 (e.g. functional derating DTCs).
"""

from __future__ import annotations

import csv
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DIAG_DIR = Path(r"C:\Projects\Diag")
V1_XLSX = DIAG_DIR / "Consolidated Diagnostic Excel_ v1 (1).xlsx"
V3_CSV = DIAG_DIR / "Consolidated Diag v3.csv"
OUT = DIAG_DIR / "Consolidated Diag v3.1.xlsx"

OUT_HEADER = [
    "#",
    "Fault Type",
    "Subject",
    "DTC Code",
    "SPN",
    "FMI",
    "Fault Name",
    "Domain",
    "ASIL",
    "Fault Level",
    "Enabling Conditions",
    "Detection Criteria",
    "Periodicity [ms]",
    "Detection Confirmation Time [ms]",
    "Healing Criteria",
    "Healing Confirmation Time [ms]",
    "Reaction",
    "VCU Request",
    "Notification",
    "Freeze Frame",
    "Variant",
    "Status",
    "Note",
]

SORT_TO_TYPE = {
    "A": "Current",
    "V": "Voltage",
    "T": "Thermal",
    "Short/Open": "Short / Open",
    "Signal": "Signal",
    "CAN": "Communication",
    "N": "Speed / Position",
    "G": "Gate Driver",
    "Memory": "Memory",
    "Component": "Component",
}


def norm(s) -> str:
    if s is None:
        return ""
    s = str(s)
    s = (
        s.replace("\u2013", "-")
        .replace("\u2014", "-")
        .replace("\u2011", "-")
        .replace("\ufffd", "-")
    )
    # V3 CSV in cp1252 mangles en-dashes to '?'; restore to '-' inside fault-name patterns
    s = re.sub(r"Over\?Temperature", "Over-Temperature", s)
    s = re.sub(r"After\?Run", "After-Run", s)
    s = re.sub(r"Read\?Back", "Read-Back", s)
    s = re.sub(r"Wear\?Out", "Wear-Out", s)
    s = re.sub(r"Linear Deration\s*\?\s*", "Linear Deration -> ", s)
    return s.strip()


def norm_dtc(s) -> str:
    return re.sub(r"\s+", " ", norm(s)).upper()


def subject_of(name: str) -> str:
    n = name.lower()
    pairs = [
        ("hvdc voltage", "HVDC Voltage"),
        ("hvdc current", "HVDC Current"),
        ("lvdc", "LVDC"),
        ("kl15", "KL15"),
        ("phase a", "Phase A"),
        ("phase b", "Phase B"),
        ("phase c", "Phase C"),
        ("phase current sum", "Phase Current Sum"),
        ("phase voltage", "Phase Voltage"),
        ("id current", "Id Current"),
        ("iq current", "Iq Current"),
        ("stator temperature", "Stator Temperature"),
        ("estimated stator", "Stator Temperature"),
        ("rotor", "Rotor Temperature"),
        ("power switch junction", "Power Switch Junction Temp"),
        ("power switch", "Power Switch Temperature"),
        ("board temperature", "Board Temperature"),
        ("coolant", "Coolant"),
        ("dc link capacitor", "DC Link Capacitor"),
        ("oil temperature", "Oil Temperature"),
        ("oil", "Oil"),
        ("active discharge", "Active Discharge"),
        ("passive discharge", "Passive Discharge"),
        ("resolver", "Resolver"),
        ("speed", "Speed"),
        ("position", "Position"),
        ("torque", "Torque"),
        ("can", "CAN"),
        ("nvm", "NVM"),
        ("watchdog", "Watchdog"),
        ("supply", "Supply"),
        ("5v analog", "5V Analog Supply"),
        ("5v digital", "5V Digital Supply"),
        ("adc reference", "ADC Reference"),
        ("parking lock", "Parking Lock"),
        ("gear", "Gear"),
        ("drive readiness", "Drive Readiness"),
        ("after-run", "After-Run"),
        ("after?run", "After-Run"),
        ("hv interlock", "HV Interlock"),
        ("interlock", "HV Interlock"),
    ]
    for k, v in pairs:
        if k in n:
            return v
    return "General"


def thermal_tier(name: str) -> int:
    """Lower number = earlier within a thermal subject."""
    n = name.lower()
    if "measurement" in n:
        return 10
    if "open circuit" in n or "short to battery" in n or "short to ground" in n or "short circuit" in n:
        return 15
    if "plausibility" in n:
        return 20
    if "functional over" in n and "derating" in n:
        return 25
    if "discharge functional over" in n or "discharge inhibited" in n:
        return 25
    if "overtemperature failure" in n or "over temperature failure" in n:
        return 40
    if "delta" in n and "junction" in n:
        return 45
    if "undertemperature" in n or "under temperature" in n:
        return 60
    if "gradient" in n:
        return 70
    return 50


def load_v1() -> tuple[list[str], list[list]]:
    wb = load_workbook(V1_XLSX, read_only=True, data_only=True)
    ws = wb["Diag matrix_V2"]
    rows = list(ws.iter_rows(min_row=3, values_only=True))
    hdr = list(rows[0])
    data = [list(r) for r in rows[1:] if any(v not in (None, "") for v in r)]
    wb.close()
    return hdr, data


def load_v3_index() -> dict[str, dict]:
    """Index V3 rows by normalized DTC and by normalized name -> {spn,fmi,dtc,name}."""
    with V3_CSV.open(newline="", encoding="cp1252") as f:
        rows = list(csv.reader(f))
    h = rows[2]
    fn_i = h.index("Fault Name")
    dt_i = h.index("DTC Code")
    spn_i = h.index("SPN")
    fmi_i = h.index("FMI")
    cov_i = h.index("Coverage")
    react_i = h.index("Reaction")
    rat_i = h.index("Rationale")
    by_dtc: dict[str, dict] = {}
    by_name: dict[str, dict] = {}
    extras: list[dict] = []
    for r in rows[3:]:
        if not r or len(r) <= fn_i:
            continue
        name = norm(r[fn_i])
        if not name:
            continue
        rec = {
            "name": name,
            "dtc": norm_dtc(r[dt_i]),
            "spn": norm(r[spn_i]),
            "fmi": norm(r[fmi_i]),
            "coverage": norm(r[cov_i]),
            "reaction": norm(r[react_i]),
            "rationale": norm(r[rat_i]),
        }
        if rec["dtc"]:
            by_dtc.setdefault(rec["dtc"], rec)
        by_name.setdefault(name.lower(), rec)
        extras.append(rec)
    return {"by_dtc": by_dtc, "by_name": by_name, "all": extras}


def dedupe_v1(hdr: list[str], data: list[list]) -> list[dict]:
    """Collapse the V1 QM+FuSa duplicate rows into one record per fault.

    Strategy: key = (Name lower, SPN, FMI). When two rows share the key, prefer the
    QM-flavour row (ASIL == 'QM'). Capture the FuSa reaction in a note suffix.
    """

    def gi(name: str) -> int:
        for i, h in enumerate(hdr):
            if h and h.strip().lower() == name.lower():
                return i
        return -1

    cols = {
        "name": gi("Name"),
        "sort": gi("Sort"),
        "desc": gi("Description"),
        "spn": gi("SPN"),
        "fmi": gi("FMI"),
        "fl": gi("Fault Level"),
        "asil": gi("ASIL"),
        "domain": gi("Domain"),
        "enab": gi("Enabling confitions"),
        "det": gi("Detection Criteria"),
        "per": gi("Periodicity [ms]"),
        "dct": gi("Detection Confirmation time [ms]"),
        "heal": gi("Healing criteria"),
        "hct": gi("Healing Confirmation time [ms]"),
        "rxn": gi("Reaction"),
        "vcu": gi("VCU Request"),
        "notif": gi("Notification"),
        "dtc": gi("DTC Code"),
        "ff": gi("Freeze frame"),
        "var": gi("Variant"),
        "note": gi("Note / Comment"),
    }

    groups: dict[tuple, list[list]] = defaultdict(list)
    order: list[tuple] = []
    for r in data:
        name = norm(r[cols["name"]])
        if not name:
            continue
        key = (name.lower(), norm(r[cols["spn"]]), norm(r[cols["fmi"]]))
        if key not in groups:
            order.append(key)
        groups[key].append(r)

    out: list[dict] = []
    for key in order:
        bucket = groups[key]
        # pick the row whose ASIL=='QM' first; else first in V1 order
        chosen = None
        other = None
        for row in bucket:
            asil = norm(row[cols["asil"]])
            if asil.upper() == "QM" and chosen is None:
                chosen = row
            elif other is None:
                other = row
        if chosen is None:
            chosen = bucket[0]
            other = bucket[1] if len(bucket) > 1 else None

        rec = {
            "name": norm(chosen[cols["name"]]),
            "sort": norm(chosen[cols["sort"]]),
            "desc": norm(chosen[cols["desc"]]),
            "spn": norm(chosen[cols["spn"]]),
            "fmi": norm(chosen[cols["fmi"]]),
            "fl": norm(chosen[cols["fl"]]),
            "asil": norm(chosen[cols["asil"]]),
            "domain": norm(chosen[cols["domain"]]),
            "enab": norm(chosen[cols["enab"]]),
            "det": norm(chosen[cols["det"]]),
            "per": norm(chosen[cols["per"]]),
            "dct": norm(chosen[cols["dct"]]),
            "heal": norm(chosen[cols["heal"]]),
            "hct": norm(chosen[cols["hct"]]),
            "rxn": norm(chosen[cols["rxn"]]),
            "vcu": norm(chosen[cols["vcu"]]),
            "notif": norm(chosen[cols["notif"]]),
            "dtc": norm_dtc(chosen[cols["dtc"]]),
            "ff": norm(chosen[cols["ff"]]),
            "var": norm(chosen[cols["var"]]),
            "note": norm(chosen[cols["note"]]),
            "status": "Original (V1)",
        }
        if other is not None:
            other_rxn = norm(other[cols["rxn"]])
            other_asil = norm(other[cols["asil"]])
            if other_rxn and other_rxn != rec["rxn"]:
                tag = f"FuSa ({other_asil}) reaction: {other_rxn}" if other_asil else f"FuSa reaction: {other_rxn}"
                rec["note"] = (rec["note"] + " | " + tag).strip(" |") if rec["note"] else tag
        out.append(rec)
    return out


def apply_overtemperature_reaction_rule(records: list[dict]) -> None:
    """For hard 'Temperature Overtemperature Failure' rows, force QM Reaction = 'ASC'.

    This honours the design rule: QM hard-OT fault must request ASC (active short
    circuit). The original V1 reaction is captured in the Note for traceability.
    Functional 'Derating Active' rows are NOT touched.
    """
    for r in records:
        n = r["name"].lower()
        is_hard_ot = (
            ("overtemperature failure" in n or "over temperature failure" in n)
            and "derating" not in n
            and "discharge" not in n
            and "delta" not in n
        )
        if not is_hard_ot:
            continue
        original = r["rxn"]
        if original.upper() == "ASC":
            continue
        r["rxn"] = "ASC"
        tag = f"Reaction set to ASC per v3.1 rule (V1 had: {original or 'blank'})"
        r["note"] = (r["note"] + " | " + tag).strip(" |") if r["note"] else tag


def add_v3_extras(records: list[dict], v3_idx: dict) -> list[dict]:
    """Append V3-only QM-relevant rows that have no V1 equivalent."""
    have_dtc = {r["dtc"] for r in records if r["dtc"]}
    have_name = {r["name"].lower() for r in records}
    added = 0
    for rec in v3_idx["all"]:
        if rec["coverage"] == "FuSa":
            continue
        if rec["dtc"] in have_dtc:
            continue
        if rec["name"].lower() in have_name:
            continue
        # only V3 rows we want: functional thermal derating + functional system DTCs
        if "Functional" not in rec["coverage"]:
            continue
        records.append(
            {
                "name": rec["name"],
                "sort": "T" if "Thermal" in rec["coverage"] or "temperature" in rec["name"].lower() else "Functional",
                "desc": rec["name"],
                "spn": rec["spn"],
                "fmi": rec["fmi"],
                "fl": "L1",
                "asil": "QM",
                "domain": "SW",
                "enab": "",
                "det": "",
                "per": "",
                "dct": "",
                "heal": "",
                "hct": "",
                "rxn": rec["reaction"] or "Linear Deration",
                "vcu": "",
                "notif": "",
                "dtc": rec["dtc"],
                "ff": "",
                "var": "Both",
                "note": rec["rationale"],
                "status": "Added (V3 AI)",
            }
        )
        added += 1
    return records


def patch_spn_fmi_from_v3(records: list[dict], v3_idx: dict) -> None:
    """Where V1 SPN/FMI is blank, fill from V3 by DTC or Name match."""
    for r in records:
        if r["spn"] and r["fmi"]:
            continue
        ref = None
        if r["dtc"]:
            ref = v3_idx["by_dtc"].get(r["dtc"])
        if ref is None:
            ref = v3_idx["by_name"].get(r["name"].lower())
        if ref:
            if not r["spn"]:
                r["spn"] = ref["spn"]
            if not r["fmi"]:
                r["fmi"] = ref["fmi"]


def fault_type_from_sort_and_subject(sort: str, subject: str, name: str) -> str:
    if subject.endswith("Temperature") or subject in {"Stator Temperature", "Rotor Temperature", "Coolant", "Oil", "Oil Temperature", "DC Link Capacitor", "Board Temperature", "Power Switch Temperature", "Power Switch Junction Temp"}:
        return "Thermal"
    if "temperature" in name.lower():
        return "Thermal"
    return SORT_TO_TYPE.get(sort.strip(), sort.strip() or "Other")


def order_records(records: list[dict]) -> list[dict]:
    """Group by (Fault Type, Subject) preserving first-seen order; within group apply thermal ordering."""
    enriched = []
    for r in records:
        subj = subject_of(r["name"])
        ft = fault_type_from_sort_and_subject(r["sort"], subj, r["name"])
        tier = thermal_tier(r["name"]) if ft == "Thermal" else 50
        enriched.append((ft, subj, tier, r))

    seen: list[tuple[str, str]] = []
    seen_set: set[tuple[str, str]] = set()
    for ft, subj, _, _ in enriched:
        if (ft, subj) not in seen_set:
            seen_set.add((ft, subj))
            seen.append((ft, subj))

    final: list[dict] = []
    for key in seen:
        members = [(t, r) for ft, sj, t, r in enriched if (ft, sj) == key]
        members.sort(key=lambda x: x[0])
        for _, r in members:
            r["_fault_type"] = key[0]
            r["_subject"] = key[1]
            final.append(r)
    return final


# ---------------------------------------------------------------------------
# Layman justification

def layman_for(rec: dict) -> str:
    n = rec["name"]
    nl = n.lower()
    rxn = rec["rxn"]

    def with_action() -> str:
        if not rxn:
            return ""
        rl = rxn.lower()
        if "linear derat" in rl:
            return "MCU smoothly cuts torque so the situation does not get worse."
        if rl == "asc":
            return "MCU shorts the three motor phases together (ASC) to stop creating heat / current immediately."
        if "asc or 6 switch open" in rl:
            return "Safety layer either shorts the three phases (ASC) or opens all six switches to stop the inverter."
        if "6 switch open" in rl:
            return "MCU opens all six power switches so the inverter no longer drives the motor."
        if "zero torque" in rl:
            return "MCU forces output torque to zero."
        if "torque limit" in rl:
            return "MCU caps the available torque while the issue is present."
        if rl.startswith("no reaction"):
            return "MCU only reports the fault; no torque action is taken."
        return f"Reaction: {rxn}."

    # specific phrases
    if "measurement failure" in nl:
        what = "the sensor reading is outside its physically possible range, so MCU can no longer trust it"
    elif "plausibility" in nl:
        what = "two sources that should agree (e.g. measured vs. estimated, or VCU vs. local) disagree by too much"
    elif "open circuit" in nl:
        what = "the wire to the sensor looks broken (open)"
    elif "short to battery" in nl:
        what = "the sensor wire is shorted to the supply rail"
    elif "short to ground" in nl:
        what = "the sensor wire is shorted to ground"
    elif "functional over" in nl and "derating" in nl:
        thing = nl.split("functional")[0].strip()
        what = f"{thing} is getting hot but still safe; this is the early warning where MCU starts derating before any safety limit is hit"
    elif "discharge functional over" in nl or "discharge inhibited" in nl:
        what = "the active-discharge resistor is heating up too much from repeated HV-off cycles, so the next active discharge is inhibited (passive discharge still works)"
    elif "overtemperature" in nl:
        thing = nl.split("overtemperature")[0].strip()
        what = f"{thing} crossed the hard safety temperature limit"
    elif "undertemperature" in nl:
        what = "the measured temperature is impossibly low — usually a sensor or wiring issue"
    elif "gradient" in nl and "temperature" in nl:
        what = "temperature is changing faster than physics allows — sensor noise / wiring issue"
    elif "overcurrent sw" in nl or ("overcurrent" in nl and "hw" not in nl):
        what = "current crossed the software overcurrent threshold"
    elif "undercurrent" in nl:
        what = "current is below the expected operating range — sensor or wiring suspect"
    elif "overcurrent hw" in nl or "over hw" in nl:
        what = "hardware comparator tripped — current crossed the hardware safety threshold"
    elif "overvoltage sw" in nl:
        what = "voltage crossed the software overvoltage threshold"
    elif "overvoltage hw" in nl:
        what = "hardware comparator tripped — voltage crossed the hardware safety threshold"
    elif "undervoltage" in nl:
        what = "voltage dropped below the safe operating threshold"
    elif "offset" in nl and "current" in nl:
        what = "with the inverter idle, the current sensor reads a non-zero offset — needs re-calibration"
    elif "current sum" in nl:
        what = "phase A + B + C currents do not sum close to zero — one of the phase sensors is wrong"
    elif "resolver excitation" in nl:
        what = "the resolver excitation signal is not as expected — rotor position sensing at risk"
    elif "resolver" in nl:
        what = "rotor-position sensor signal is bad"
    elif "can" in nl and ("timeout" in nl or "missing" in nl):
        what = "expected CAN message did not arrive in time"
    elif "nvm" in nl:
        what = "non-volatile memory check failed (CRC / wear) — calibrations may be unreliable"
    elif "watchdog" in nl:
        what = "the safety watchdog did not get its expected pulse — software may be hung"
    elif "interlock" in nl:
        what = "the HV interlock loop is broken — assume HV is unsafe to energize"
    elif "measurement fault" in nl or "measurement failure" in nl:
        what = "the sensor reading is outside its physically possible range, so MCU can no longer trust it"
    elif "locked rotor" in nl:
        what = "rotor is commanded but not turning (mechanical lock or sensor disagreement)"
    elif "parking lock" in nl:
        what = "the parking-lock state from VCU does not match what the MCU sees"
    elif "after-run" in nl:
        what = "the after-run shutdown sequence did not complete in time"
    else:
        what = n.lower()

    sentence = f"What it means: {what}."
    action = with_action()
    if action:
        sentence += " " + action
    return sentence


# ---------------------------------------------------------------------------
# Build workbook

def build() -> None:
    hdr1, data1 = load_v1()
    v1_records = dedupe_v1(hdr1, data1)
    v3 = load_v3_index()
    patch_spn_fmi_from_v3(v1_records, v3)
    apply_overtemperature_reaction_rule(v1_records)
    records = add_v3_extras(v1_records, v3)
    records = order_records(records)

    wb = Workbook()
    ws = wb.active
    ws.title = "Master v3.1 QM"

    title = "HANDE MCU - Diagnostic Matrix v3.1 (QM)"
    sub = (
        "Source: V1 sheet 'Diag matrix_V2' (deduplicated). V3 used only to fill SPN/FMI gaps "
        "and to add V3-AI functional faults. Status column tags origin of each row."
    )
    ws.append([title] + [""] * (len(OUT_HEADER) - 1))
    ws.append([sub] + [""] * (len(OUT_HEADER) - 1))
    ws.append(OUT_HEADER)

    status_fill = {
        "Original (V1)": PatternFill("solid", fgColor="E8F1FB"),
        "Added (V3 AI)": PatternFill("solid", fgColor="FFF4D6"),
        "Added (Customer)": PatternFill("solid", fgColor="E8F8E1"),
        "Added (new in v3.1)": PatternFill("solid", fgColor="F4E1F8"),
    }

    status_idx = OUT_HEADER.index("Status")

    for idx, r in enumerate(records, start=1):
        row_vals = [
            idx,
            r.get("_fault_type", ""),
            r.get("_subject", ""),
            r["dtc"],
            r["spn"],
            r["fmi"],
            r["name"],
            r["domain"],
            r["asil"],
            r["fl"],
            r["enab"],
            r["det"],
            r["per"],
            r["dct"],
            r["heal"],
            r["hct"],
            r["rxn"],
            r["vcu"],
            r["notif"],
            r["ff"],
            r["var"],
            r["status"],
            r["note"],
        ]
        ws.append(row_vals)
        fill = status_fill.get(r["status"])
        if fill:
            ws.cell(row=ws.max_row, column=status_idx + 1).fill = fill

    # styling
    for r in (1, 2):
        for c in ws[r]:
            c.font = Font(bold=True)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F3864")
    for c in ws[3]:
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 36
    ws.row_dimensions[3].height = 32

    widths = {
        "#": 5, "Fault Type": 14, "Subject": 22, "DTC Code": 12, "SPN": 8, "FMI": 6,
        "Fault Name": 46, "Domain": 8, "ASIL": 8, "Fault Level": 10,
        "Enabling Conditions": 22, "Detection Criteria": 30, "Periodicity [ms]": 10,
        "Detection Confirmation Time [ms]": 12, "Healing Criteria": 30,
        "Healing Confirmation Time [ms]": 12, "Reaction": 24, "VCU Request": 18,
        "Notification": 22, "Freeze Frame": 18, "Variant": 10, "Status": 18, "Note": 32,
    }
    for i, h in enumerate(OUT_HEADER, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(h, 14)

    for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "D4"

    # Justifications sheet
    wj = wb.create_sheet("Justifications")
    j_hdr = ["#", "DTC Code", "Fault Name", "What this means (layman terms)", "Status"]
    wj.append(j_hdr)
    for c in wj[1]:
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for n, r in enumerate(records, start=1):
        wj.append([n, r["dtc"], r["name"], layman_for(r), r["status"]])
        fill = status_fill.get(r["status"])
        if fill:
            wj.cell(row=wj.max_row, column=5).fill = fill
    wj.column_dimensions["A"].width = 5
    wj.column_dimensions["B"].width = 12
    wj.column_dimensions["C"].width = 46
    wj.column_dimensions["D"].width = 95
    wj.column_dimensions["E"].width = 18
    wj.row_dimensions[1].height = 28
    for row in wj.iter_rows(min_row=2, max_row=wj.max_row):
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical="top")
    wj.freeze_panes = "B2"

    # Legend sheet
    wl = wb.create_sheet("Legend", 0)
    wl.append(["Consolidated Diag v3.1"])
    wl.cell(1, 1).font = Font(bold=True, size=14)
    wl.append([])
    wl.append(["Sheet", "Purpose"])
    for c in wl[3]:
        c.font = header_font
        c.fill = header_fill
    wl.append(["Master v3.1 QM", "Full QM diagnostic matrix"])
    wl.append(["Justifications", "Plain-English explanation per fault"])
    wl.append([])
    wl.append(["Status value", "Meaning", "Colour"])
    for c in wl[wl.max_row]:
        c.font = header_font
        c.fill = header_fill
    legend_rows = [
        ("Original (V1)", "Row taken from V1 sheet 'Diag matrix_V2' (deduplicated)."),
        ("Added (V3 AI)", "Functional / derating row added by the V3 AI consolidation."),
        ("Added (Customer)", "Row brought in from the customer DTC list."),
        ("Added (new in v3.1)", "Row newly added in this revision (gap-fill)."),
    ]
    for s, m in legend_rows:
        wl.append([s, m, ""])
        wl.cell(wl.max_row, 3).fill = status_fill[s]
    wl.column_dimensions["A"].width = 22
    wl.column_dimensions["B"].width = 70
    wl.column_dimensions["C"].width = 12

    wb.save(OUT)
    counts: dict[str, int] = defaultdict(int)
    for r in records:
        counts[r["status"]] += 1
    print(f"Wrote {OUT}")
    print(f"  rows: {len(records)}  -> {dict(counts)}")


if __name__ == "__main__":
    build()
