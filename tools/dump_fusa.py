import openpyxl
p = r"C:\Projects\Diag\Consolidated Diag v3.2 Review .xlsx"
wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
ws = wb["Diag matrix_Fusa"]
hdr = list(next(ws.iter_rows(min_row=3, max_row=3, values_only=True)))
print("FUSA HEADER:", hdr)


def gi(name):
    for i, h in enumerate(hdr):
        if h and str(h).strip().lower() == name.lower():
            return i
    return -1


nm = gi("Name")
spn = gi("SPN")
fmi = gi("FMI")
dtc = gi("DTC Code")
rxn = gi("Reaction")
fl = gi("Fault Level")
asil = gi("ASIL")

print()
print("FUSA rows:")
for i, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
    if not row or not row[nm]:
        continue
    name = str(row[nm] or "")[:55]
    d = str(row[dtc] or "")
    s = str(row[spn] or "")
    f = str(row[fmi] or "")
    flv = str(row[fl] or "")
    a = str(row[asil] or "")
    r = str(row[rxn] or "")[:40]
    print(
        f"r{i:>3} | {d:<11}| SPN={s:<7}| FMI={f:<3}| FL={flv:<3}| ASIL={a:<6}| {name:<55}| rxn={r}"
    )
wb.close()
